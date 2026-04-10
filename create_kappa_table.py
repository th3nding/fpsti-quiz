import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import math

# 分析结果数据
dimensions = [
    "技能修改程度",
    "地图物件来源", 
    "机制创新",
    "视觉创新",
    "叙事创新"
]

kappa_values = [0.8806, 0.9152, 0.9091, 0.9206, 0.9600]
agreement_rates = [91.00, 94.00, 93.00, 94.00, 97.00]
n = 100  # 样本数

# 计算统计量
# Kappa 系数的标准误公式：SE(κ) = sqrt(κ(1-κ) / (n * (1-pe)^2))
# 简化计算：使用近似公式 SE ≈ sqrt((1-κ) / (n * κ))
# z 值 = κ / SE
# 95% CI = κ ± 1.96 * SE

results = []
for i, dim in enumerate(dimensions):
    kappa = kappa_values[i]
    
    # 计算标准误 (使用简化公式)
    # 更准确的公式：SE(κ) ≈ sqrt(κ(1-κ) / n) / (1 - Pe)
    # 这里 Pe 约等于 0.25 (根据之前的计算)
    pe_approx = 0.25
    se = math.sqrt(kappa * (1 - kappa) / n) / (1 - pe_approx)
    
    # z 值
    z = kappa / se
    
    # p 值 (使用近似，z > 3 时 p < 0.001)
    if z > 3.29:
        p_value = 0.001
        p_sign = "***"
    elif z > 2.58:
        p_value = 0.01
        p_sign = "**"
    elif z > 1.96:
        p_value = 0.05
        p_sign = "*"
    else:
        p_value = 0.1
        p_sign = ""
    
    # 95% 置信区间
    ci_lower = kappa - 1.96 * se
    ci_upper = kappa + 1.96 * se
    
    results.append({
        "dimension": dim,
        "kappa": kappa,
        "se": se,
        "z": z,
        "p_value": p_value,
        "p_sign": p_sign,
        "ci_lower": ci_lower,
        "ci_upper": ci_upper
    })

# 创建 workbook
wb = Workbook()
ws = wb.active
ws.title = "Kappa 系数结果"

# 设置样式
title_font = Font(bold=True, size=12, name="宋体")
header_font = Font(bold=True, size=10, name="宋体")
normal_font = Font(size=10, name="宋体")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
thick_border_top = Border(
    top=Side(style='thick'),
    bottom=Side(style='thin'),
    left=Side(style='thin'),
    right=Side(style='thin')
)
thick_border_bottom = Border(
    top=Side(style='thin'),
    bottom=Side(style='thick'),
    left=Side(style='thin'),
    right=Side(style='thin')
)

# 添加标题
ws.merge_cells('A1:G1')
ws['A1'] = "系数结果表格"
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center')

# 添加表头
headers = ["名称", "κ值", "标准误 (假定原假设)", "z 值", "p 值", "标准误", "95% CI"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=header)
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thick_border_top

# 添加数据
for i, result in enumerate(results, 3):
    # 名称
    ws.cell(row=i, column=1, value=f"{result['dimension']}1 & {result['dimension']}2").font = normal_font
    
    # κ值
    ws.cell(row=i, column=2, value=round(result['kappa'], 3)).font = normal_font
    ws.cell(row=i, column=2).alignment = Alignment(horizontal='center')
    
    # 标准误 (假定原假设)
    se_null = round(result['kappa'] * (1 - result['kappa']) / 100, 3)
    ws.cell(row=i, column=3, value=se_null).font = normal_font
    ws.cell(row=i, column=3).alignment = Alignment(horizontal='center')
    
    # z 值
    ws.cell(row=i, column=4, value=round(result['z'], 3)).font = normal_font
    ws.cell(row=i, column=4).alignment = Alignment(horizontal='center')
    
    # p 值
    p_display = f"{result['p_value']:.3f}{result['p_sign']}" if result['p_value'] >= 0.001 else f"<0.001{result['p_sign']}"
    ws.cell(row=i, column=5, value=p_display).font = normal_font
    ws.cell(row=i, column=5).alignment = Alignment(horizontal='center')
    
    # 标准误
    ws.cell(row=i, column=6, value=round(result['se'], 3)).font = normal_font
    ws.cell(row=i, column=6).alignment = Alignment(horizontal='center')
    
    # 95% CI
    ci_text = f"{result['ci_lower']:.3f} ~ {result['ci_upper']:.3f}"
    ws.cell(row=i, column=7, value=ci_text).font = normal_font
    ws.cell(row=i, column=7).alignment = Alignment(horizontal='center')
    
    # 设置边框
    for col in range(1, 8):
        if i == len(results) + 2:
            ws.cell(row=i, column=col).border = thick_border_bottom
        else:
            ws.cell(row=i, column=col).border = thin_border

# 添加注释
ws.merge_cells('A8:G8')
ws['A8'] = "* p<0.05  ** p<0.01  *** p<0.001"
ws['A8'].font = Font(size=9, name="宋体", italic=True)

# 设置列宽
column_widths = [28, 10, 22, 12, 12, 12, 22]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# 设置行高
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 25
for i in range(3, 8):
    ws.row_dimensions[i].height = 22

# 创建汇总表
ws2 = wb.create_sheet(title="汇总表")
ws2['A1'] = "编码员一致性分析汇总"
ws2['A1'].font = Font(bold=True, size=14, name="宋体")

summary_headers = ["维度", "一致率 (%)", "Kappa 系数", "信度等级"]
for col, header in enumerate(summary_headers, 1):
    cell = ws2.cell(row=3, column=col, value=header)
    cell.font = Font(bold=True, name="宋体")
    cell.alignment = Alignment(horizontal='center')

for i, (dim, rate, kappa) in enumerate(zip(dimensions, agreement_rates, kappa_values), 4):
    ws2.cell(row=i, column=1, value=dim).font = Font(name="宋体")
    ws2.cell(row=i, column=2, value=f"{rate}%").font = Font(name="宋体")
    ws2.cell(row=i, column=2).alignment = Alignment(horizontal='center')
    ws2.cell(row=i, column=3, value=kappa).font = Font(name="宋体")
    ws2.cell(row=i, column=3).alignment = Alignment(horizontal='center')
    
    # 信度等级
    if kappa > 0.9:
        grade = "几乎完全一致 ⭐"
    else:
        grade = "几乎完全一致"
    ws2.cell(row=i, column=4, value=grade).font = Font(name="宋体")

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 18

# 保存
output_path = r"C:\Users\shenyue05\Desktop\毕业论文\模组（质性）\Kappa 系数结果表格.xlsx"
wb.save(output_path)
print(f"✓ 结果已保存到：{output_path}")
print(f"\n表格包含:")
print(f"  1. Kappa 系数结果 - 学术论文格式")
print(f"  2. 汇总表 - 简洁版汇总")
