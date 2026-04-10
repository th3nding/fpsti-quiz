import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

dimensions = ["Skill Modification", "Map Object Source", "Mechanism Innovation", "Visual Innovation", "Narrative Innovation"]
agreement_rates = [91.00, 94.00, 93.00, 94.00, 97.00]
agreements = [91, 94, 93, 94, 97]
disagreements = [9, 6, 7, 6, 3]
kappa_values = [0.8806, 0.9152, 0.9091, 0.9206, 0.9600]

wb = Workbook()
ws = wb.active
ws.title = "Coder Reliability"

title_font = Font(bold=True, size=16, color="FFFFFF")
header_font = Font(bold=True, size=12)
title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

ws.merge_cells('A1:E1')
ws['A1'] = "Overwatch Map Workshop - Coder Reliability Analysis"
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

ws.merge_cells('A2:E2')
ws['A2'] = "Total: 100 samples | Overall Agreement: 93.80% | All Kappa > 0.88"
ws['A2'].font = Font(bold=True, size=11)
ws['A2'].alignment = Alignment(horizontal='center')
ws.row_dimensions[2].height = 30

headers = ["Dimension", "Agreement %", "Agreements", "Disagreements", "Kappa"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for i, dim in enumerate(dimensions, 4):
    ws.cell(row=i, column=1, value=dim).border = thin_border
    ws.cell(row=i, column=2, value=agreement_rates[i-4]).border = thin_border
    ws.cell(row=i, column=3, value=agreements[i-4]).border = thin_border
    ws.cell(row=i, column=4, value=disagreements[i-4]).border = thin_border
    ws.cell(row=i, column=5, value=kappa_values[i-4]).border = thin_border

ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 12

chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Agreement Rate by Dimension"
chart1.y_axis.title = "Agreement Rate (%)"
chart1.x_axis.title = "Dimension"
chart1.legend = None
data1 = Reference(ws, min_col=2, min_row=3, max_row=8)
cats1 = Reference(ws, min_col=1, min_row=4, max_row=8)
chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.shape = 4
chart1.anchor = "G3"

chart2 = BarChart()
chart2.type = "col"
chart2.style = 10
chart2.title = "Cohen's Kappa Coefficient"
chart2.y_axis.title = "Kappa"
chart2.x_axis.title = "Dimension"
chart2.legend = None
data2 = Reference(ws, min_col=5, min_row=3, max_row=8)
cats2 = Reference(ws, min_col=1, min_row=4, max_row=8)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.shape = 4
chart2.anchor = "G20"

ws2 = wb.create_sheet(title="Disagreements")
detail_headers = ["Dimension", "Row", "Work", "Coder1", "Coder2", "Diff"]
for col, header in enumerate(detail_headers, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

disagreement_details = {
    "Skill Modification": [(3, "Flying Cat", 1, 2), (6, "Hex Clash", 4, 5), (22, "Graveyard", 2, 1), (24, "Tower Defense", 4, 3), (37, "Fight Landlord", 5, 4), (46, "Cliff Brawl", 4, 5), (64, "Sim Ladder", 1, 2), (67, "9sqm Genji", 1, 2), (96, "Rat Revenge", 4, 2)],
    "Map Object Source": [(9, "Lijiang 1v1", 1, 2), (56, "Pizza Town", 5, 2), (59, "Genji Practice", 1, 3), (76, "Zombie World", 4, 2), (84, "Resource Rush", 3, 1), (87, "Hammer Zombie", 3, 1)],
    "Mechanism Innovation": [(9, "Lijiang 1v1", 3, 2), (27, "Hamster Quiz", 4, 5), (34, "Zombie Tide", 4, 3), (46, "Cliff Brawl", 2, 3), (67, "9sqm Genji", 2, 1), (91, "Pest Control", 5, 2), (100, "Protect Pres", 4, 2)],
    "Visual Innovation": [(9, "Lijiang 1v1", 2, 1), (20, "Clover Abyss", 5, 3), (49, "Tower Sim", 4, 2), (60, "Ghosts Moon", 3, 5), (71, "BOSS Arena", 4, 5), (80, "Zombie Infection", 3, 5)],
    "Narrative Innovation": [(27, "Hamster Quiz", 2, 1), (51, "Sekiro PVE", 4, 2), (100, "Protect Pres", 5, 3)]
}

row = 2
for dim, details in disagreement_details.items():
    for detail in details:
        ws2.cell(row=row, column=1, value=dim)
        ws2.cell(row=row, column=2, value=detail[0])
        ws2.cell(row=row, column=3, value=detail[1])
        ws2.cell(row=row, column=4, value=detail[2])
        ws2.cell(row=row, column=5, value=detail[3])
        ws2.cell(row=row, column=6, value=abs(detail[2] - detail[3]))
        row += 1

ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 6
ws2.column_dimensions['C'].width = 20
ws2.column_dimensions['D'].width = 8
ws2.column_dimensions['E'].width = 8
ws2.column_dimensions['F'].width = 6

ws3 = wb.create_sheet(title="Summary")
ws3['A1'] = "CODER RELIABILITY ANALYSIS SUMMARY"
ws3['A1'].font = Font(bold=True, size=14)
ws3['A3'] = "Overall Results:"
ws3['A3'].font = Font(bold=True)
ws3['A4'] = "- Total samples: 100 Overwatch map workshop mods"
ws3['A5'] = "- Overall agreement rate: 93.80% (469/500 codes)"
ws3['A6'] = "- All Kappa coefficients > 0.88 (Almost Perfect agreement)"
ws3['A8'] = "By Dimension:"
ws3['A8'].font = Font(bold=True)
ws3['A9'] = "1. Narrative Innovation: 97.00% (Kappa=0.96) - BEST"
ws3['A10'] = "2. Map Object Source: 94.00% (Kappa=0.92)"
ws3['A11'] = "3. Visual Innovation: 94.00% (Kappa=0.92)"
ws3['A12'] = "4. Mechanism Innovation: 93.00% (Kappa=0.91)"
ws3['A13'] = "5. Skill Modification: 91.00% (Kappa=0.88) - Needs refinement"
ws3['A15'] = "Conclusion: Excellent coding reliability. All dimensions meet academic standards (Kappa > 0.85)."
ws3['A16'] = "Total disagreements: 31 cases across all dimensions."
ws3.column_dimensions['A'].width = 70

ws.add_chart(chart1, "G3")
ws.add_chart(chart2, "G20")

output_path = r"C:\Users\shenyue05\Desktop\毕业论文\模组（质性）\Coder_Reliability_Analysis.xlsx"
wb.save(output_path)
print(f"Analysis saved to: {output_path}")
print(f"\nWorksheets:")
print(f"  1. Coder Reliability - Summary data + charts")
print(f"  2. Disagreements - 31 disagreement cases")
print(f"  3. Summary - Key findings")
